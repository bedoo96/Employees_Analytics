"""
Report Generator Module
Generates comprehensive attendance reports in Excel and PDF formats
with insights, comments, and recommendations for HR follow-up
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import io
from PIL import Image
import plotly.graph_objects as go
import plotly.express as px


class ReportGenerator:
    """
    Comprehensive report generator for attendance analytics
    Supports Excel and PDF exports with insights and recommendations
    """
    
    def __init__(self, analyzer):
        """
        Initialize report generator with attendance analyzer
        
        Args:
            analyzer: AttendanceAnalyzer instance with loaded data
        """
        self.analyzer = analyzer
        self.df = analyzer.df
        self.month_year = analyzer.month_year
        self.insights = self._generate_insights()
        
    def _generate_insights(self):
        """Generate automated insights from the data"""
        insights = {
            'summary': {},
            'concerns': [],
            'recommendations': [],
            'highlights': []
        }
        
        # Summary statistics
        insights['summary'] = {
            'total_employees': len(self.df),
            'avg_working_hours': self.df['Regular(H)'].mean() if 'Regular(H)' in self.df.columns else 0,
            'total_late_minutes': self.df['Late In(M)'].sum() if 'Late In(M)' in self.df.columns else 0,
            'total_overtime': self.df['Normal OT(H)'].sum() if 'Normal OT(H)' in self.df.columns else 0,
            'total_absences': self.df['Absence(H)'].sum() if 'Absence(H)' in self.df.columns else 0,
        }
        
        # Identify concerns
        if 'Late In(M)' in self.df.columns:
            late_employees = len(self.df[self.df['Late In(M)'] > 60])
            if late_employees > 0:
                insights['concerns'].append(
                    f"{late_employees} employees have excessive late arrivals (>60 min)"
                )
        
        if 'Normal OT(H)' in self.df.columns:
            high_ot = len(self.df[self.df['Normal OT(H)'] > 30])
            if high_ot > 0:
                insights['concerns'].append(
                    f"{high_ot} employees have high overtime (>30 hours) - potential burnout risk"
                )
        
        if 'Absence(H)' in self.df.columns:
            high_absence = len(self.df[self.df['Absence(H)'] > 16])
            if high_absence > 0:
                insights['concerns'].append(
                    f"{high_absence} employees have significant absences (>2 days)"
                )
        
        # Generate recommendations
        avg_late = self.df['Late In(M)'].mean() if 'Late In(M)' in self.df.columns else 0
        if avg_late > 20:
            insights['recommendations'].append(
                "Consider reviewing start time policies or flexible work arrangements"
            )
        
        avg_ot = self.df['Normal OT(H)'].mean() if 'Normal OT(H)' in self.df.columns else 0
        if avg_ot > 10:
            insights['recommendations'].append(
                "High average overtime detected - review workload distribution and consider additional staffing"
            )
        
        # Highlights
        if 'Regular(H)' in self.df.columns:
            top_performer = self.df.nlargest(1, 'Regular(H)')
            if not top_performer.empty:
                insights['highlights'].append(
                    f"Top performer: {top_performer['First Name'].values[0]} with {top_performer['Regular(H)'].values[0]:.1f} working hours"
                )
        
        return insights
    
    def generate_excel_report(self, output_path='attendance_report.xlsx'):
        """
        Generate comprehensive Excel report with multiple sheets and insights
        
        Args:
            output_path: Path to save the Excel file
            
        Returns:
            str: Path to the generated file
        """
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Executive Summary
            self._create_executive_summary_sheet(writer)
            
            # Sheet 2: Detailed Employee Data
            self._create_detailed_data_sheet(writer)
            
            # Sheet 3: Late Arrivals Analysis
            self._create_late_arrivals_sheet(writer)
            
            # Sheet 4: Overtime Analysis
            self._create_overtime_sheet(writer)
            
            # Sheet 5: Leave Analysis
            self._create_leave_analysis_sheet(writer)
            
            # Sheet 6: Department Summary
            self._create_department_summary_sheet(writer)
            
            # Sheet 7: Action Items
            self._create_action_items_sheet(writer)
        
        # Add formatting and comments
        self._format_excel_workbook(output_path)
        
        return output_path
    
    def _create_executive_summary_sheet(self, writer):
        """Create executive summary sheet"""
        summary_data = {
            'Metric': [
                'Report Period',
                'Total Employees',
                'Average Working Hours',
                'Total Late Minutes',
                'Total Overtime Hours',
                'Total Absence Hours',
                'Employees Needing Attention',
            ],
            'Value': [
                self.month_year,
                self.insights['summary']['total_employees'],
                f"{self.insights['summary']['avg_working_hours']:.2f}",
                f"{self.insights['summary']['total_late_minutes']:.0f}",
                f"{self.insights['summary']['total_overtime']:.2f}",
                f"{self.insights['summary']['total_absences']:.2f}",
                len(self.insights['concerns']),
            ],
            'Status': [
                'Info',
                'Info',
                '✓' if self.insights['summary']['avg_working_hours'] >= 160 else '⚠',
                '⚠' if self.insights['summary']['total_late_minutes'] > 500 else '✓',
                '⚠' if self.insights['summary']['total_overtime'] > 200 else '✓',
                '⚠' if self.insights['summary']['total_absences'] > 100 else '✓',
                '⚠' if len(self.insights['concerns']) > 0 else '✓',
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # Add insights section
        insights_start_row = len(df_summary) + 3
        worksheet = writer.sheets['Executive Summary']
        
        worksheet.cell(row=insights_start_row, column=1, value='KEY CONCERNS:')
        for i, concern in enumerate(self.insights['concerns'], start=1):
            worksheet.cell(row=insights_start_row + i, column=1, value=f"• {concern}")
        
        rec_start_row = insights_start_row + len(self.insights['concerns']) + 2
        worksheet.cell(row=rec_start_row, column=1, value='RECOMMENDATIONS:')
        for i, rec in enumerate(self.insights['recommendations'], start=1):
            worksheet.cell(row=rec_start_row + i, column=1, value=f"• {rec}")
    
    def _create_detailed_data_sheet(self, writer):
        """Create detailed employee data sheet with comments"""
        detailed_df = self.df.copy()
        
        # Add insight column
        detailed_df['HR_Notes'] = detailed_df.apply(self._generate_employee_note, axis=1)
        detailed_df['Priority'] = detailed_df.apply(self._calculate_priority, axis=1)
        
        detailed_df.to_excel(writer, sheet_name='Employee Details', index=False)
    
    def _create_late_arrivals_sheet(self, writer):
        """Create late arrivals analysis sheet"""
        if 'Late In(M)' in self.df.columns:
            late_df = self.df[self.df['Late In(M)'] > 0].copy()
            late_df = late_df.sort_values('Late In(M)', ascending=False)
            
            late_df['Days_Late_Estimate'] = (late_df['Late In(M)'] / 30).apply(np.ceil)
            late_df['Severity'] = late_df['Late In(M)'].apply(
                lambda x: 'High' if x > 100 else ('Medium' if x > 50 else 'Low')
            )
            late_df['Follow_Up_Required'] = late_df['Late In(M)'].apply(
                lambda x: 'Yes - Immediate' if x > 100 else ('Yes' if x > 50 else 'Monitor')
            )
            
            cols = ['Employee ID', 'First Name', 'Department', 'Late In(M)', 
                   'Days_Late_Estimate', 'Severity', 'Follow_Up_Required']
            late_df[cols].to_excel(writer, sheet_name='Late Arrivals', index=False)
    
    def _create_overtime_sheet(self, writer):
        """Create overtime analysis sheet"""
        ot_cols = ['Normal OT(H)', 'Weekend OT(H)', 'Holiday OT(H)']
        available_ot = [col for col in ot_cols if col in self.df.columns]
        
        if available_ot:
            ot_df = self.df.copy()
            ot_df['Total_OT'] = ot_df[available_ot].sum(axis=1)
            ot_df = ot_df[ot_df['Total_OT'] > 0].sort_values('Total_OT', ascending=False)
            
            ot_df['OT_Level'] = ot_df['Total_OT'].apply(
                lambda x: 'Excessive (>30h)' if x > 30 else ('High (20-30h)' if x > 20 else ('Moderate (10-20h)' if x > 10 else 'Normal (<10h)'))
            )
            ot_df['Action_Required'] = ot_df['Total_OT'].apply(
                lambda x: 'Review workload & consider support' if x > 30 else ('Monitor' if x > 20 else 'None')
            )
            
            cols = ['Employee ID', 'First Name', 'Department'] + available_ot + ['Total_OT', 'OT_Level', 'Action_Required']
            ot_df[cols].to_excel(writer, sheet_name='Overtime Analysis', index=False)
    
    def _create_leave_analysis_sheet(self, writer):
        """Create leave analysis sheet"""
        leave_cols = [col for col in self.df.columns if 'Leave' in col]
        
        if leave_cols:
            leave_df = self.df.copy()
            leave_df['Total_Leave'] = leave_df[leave_cols].sum(axis=1)
            leave_df = leave_df[leave_df['Total_Leave'] > 0].sort_values('Total_Leave', ascending=False)
            
            cols = ['Employee ID', 'First Name', 'Department'] + leave_cols + ['Total_Leave']
            leave_df[cols].to_excel(writer, sheet_name='Leave Analysis', index=False)
    
    def _create_department_summary_sheet(self, writer):
        """Create department summary sheet"""
        if 'Department' not in self.df.columns:
            return
        
        agg_dict = {
            'Employee ID': 'count',
        }
        
        if 'Regular(H)' in self.df.columns:
            agg_dict['Regular(H)'] = ['sum', 'mean']
        if 'Late In(M)' in self.df.columns:
            agg_dict['Late In(M)'] = 'sum'
        if 'Normal OT(H)' in self.df.columns:
            agg_dict['Normal OT(H)'] = 'sum'
        
        dept_summary = self.df.groupby('Department').agg(agg_dict).reset_index()
        dept_summary.columns = ['_'.join(col).strip('_') for col in dept_summary.columns.values]
        
        dept_summary.to_excel(writer, sheet_name='Department Summary', index=False)
    
    def _create_action_items_sheet(self, writer):
        """Create action items for HR follow-up"""
        action_items = []
        
        # Late arrivals requiring action
        if 'Late In(M)' in self.df.columns:
            late_critical = self.df[self.df['Late In(M)'] > 100]
            for _, row in late_critical.iterrows():
                action_items.append({
                    'Priority': 'High',
                    'Employee': f"{row['First Name']} (ID: {row['Employee ID']})",
                    'Issue': 'Excessive Late Arrivals',
                    'Details': f"{row['Late In(M)']:.0f} minutes total",
                    'Recommended_Action': 'Schedule counseling meeting',
                    'Timeline': 'This week'
                })
        
        # Excessive overtime
        if 'Normal OT(H)' in self.df.columns:
            ot_critical = self.df[self.df['Normal OT(H)'] > 30]
            for _, row in ot_critical.iterrows():
                action_items.append({
                    'Priority': 'High',
                    'Employee': f"{row['First Name']} (ID: {row['Employee ID']})",
                    'Issue': 'Excessive Overtime',
                    'Details': f"{row['Normal OT(H)']:.1f} hours",
                    'Recommended_Action': 'Review workload, check burnout risk',
                    'Timeline': 'This week'
                })
        
        # High absences
        if 'Absence(H)' in self.df.columns:
            absence_critical = self.df[self.df['Absence(H)'] > 16]
            for _, row in absence_critical.iterrows():
                action_items.append({
                    'Priority': 'Medium',
                    'Employee': f"{row['First Name']} (ID: {row['Employee ID']})",
                    'Issue': 'High Absence Rate',
                    'Details': f"{row['Absence(H)']:.1f} hours",
                    'Recommended_Action': 'Wellness check, review circumstances',
                    'Timeline': 'Within 2 weeks'
                })
        
        if action_items:
            df_actions = pd.DataFrame(action_items)
            df_actions = df_actions.sort_values('Priority')
            df_actions.to_excel(writer, sheet_name='Action Items', index=False)
        else:
            # Create empty sheet with message
            df_actions = pd.DataFrame({'Message': ['No critical action items - all metrics within acceptable ranges']})
            df_actions.to_excel(writer, sheet_name='Action Items', index=False)
    
    def _generate_employee_note(self, row):
        """Generate HR note for an employee"""
        notes = []
        
        if 'Late In(M)' in row and row['Late In(M)'] > 60:
            notes.append(f"Late arrivals: {row['Late In(M)']:.0f}min")
        
        if 'Normal OT(H)' in row and row['Normal OT(H)'] > 20:
            notes.append(f"High OT: {row['Normal OT(H)']:.1f}h")
        
        if 'Absence(H)' in row and row['Absence(H)'] > 16:
            notes.append(f"High absence: {row['Absence(H)']:.1f}h")
        
        return ' | '.join(notes) if notes else 'No issues'
    
    def _calculate_priority(self, row):
        """Calculate priority level for employee"""
        score = 0
        
        if 'Late In(M)' in row and row['Late In(M)'] > 100:
            score += 3
        elif 'Late In(M)' in row and row['Late In(M)'] > 50:
            score += 1
        
        if 'Normal OT(H)' in row and row['Normal OT(H)'] > 30:
            score += 3
        elif 'Normal OT(H)' in row and row['Normal OT(H)'] > 20:
            score += 1
        
        if 'Absence(H)' in row and row['Absence(H)'] > 16:
            score += 2
        
        if score >= 5:
            return 'High'
        elif score >= 2:
            return 'Medium'
        else:
            return 'Low'
    
    def _format_excel_workbook(self, file_path):
        """Apply formatting and add comments to Excel workbook"""
        wb = load_workbook(file_path)
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add conditional formatting for status/priority columns
            if 'Status' in [cell.value for cell in ws[1]]:
                status_col = [cell.value for cell in ws[1]].index('Status') + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=status_col)
                    if '⚠' in str(cell.value):
                        cell.fill = warning_fill
                    elif '✓' in str(cell.value):
                        cell.fill = good_fill
        
        wb.save(file_path)
    
    def generate_pdf_report(self, output_path='attendance_report.pdf'):
        """
        Generate comprehensive PDF report for HR
        
        Args:
            output_path: Path to save the PDF file
            
        Returns:
            str: Path to the generated file
        """
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        story.append(Paragraph(f"Attendance Report - {self.month_year}", title_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        summary_data = [
            ['Metric', 'Value', 'Status'],
            ['Total Employees', str(self.insights['summary']['total_employees']), 'Info'],
            ['Avg Working Hours', f"{self.insights['summary']['avg_working_hours']:.2f}", 
             '✓' if self.insights['summary']['avg_working_hours'] >= 160 else '⚠'],
            ['Total Late Minutes', f"{self.insights['summary']['total_late_minutes']:.0f}",
             '⚠' if self.insights['summary']['total_late_minutes'] > 500 else '✓'],
            ['Total Overtime', f"{self.insights['summary']['total_overtime']:.2f}h",
             '⚠' if self.insights['summary']['total_overtime'] > 200 else '✓'],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 1.5*inch, 1*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Key Concerns
        if self.insights['concerns']:
            story.append(Paragraph("Key Concerns", heading_style))
            for concern in self.insights['concerns']:
                story.append(Paragraph(f"• {concern}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        # Recommendations
        if self.insights['recommendations']:
            story.append(Paragraph("Recommendations", heading_style))
            for rec in self.insights['recommendations']:
                story.append(Paragraph(f"• {rec}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
        
        story.append(PageBreak())
        
        # Late Arrivals Section
        if 'Late In(M)' in self.df.columns:
            story.append(Paragraph("Late Arrivals Analysis", heading_style))
            late_df = self.df[self.df['Late In(M)'] > 50].nlargest(10, 'Late In(M)')
            
            if not late_df.empty:
                late_data = [['Employee', 'Department', 'Late Minutes', 'Action']]
                for _, row in late_df.iterrows():
                    action = 'Immediate' if row['Late In(M)'] > 100 else 'Follow-up'
                    late_data.append([
                        row['First Name'],
                        row.get('Department', 'N/A'),
                        f"{row['Late In(M)']:.0f}",
                        action
                    ])
                
                late_table = Table(late_data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.3*inch])
                late_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(late_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Overtime Section
        if 'Normal OT(H)' in self.df.columns:
            story.append(Paragraph("Overtime Analysis", heading_style))
            ot_df = self.df[self.df['Normal OT(H)'] > 20].nlargest(10, 'Normal OT(H)')
            
            if not ot_df.empty:
                ot_data = [['Employee', 'Department', 'OT Hours', 'Level']]
                for _, row in ot_df.iterrows():
                    level = 'High' if row['Normal OT(H)'] > 30 else 'Moderate'
                    ot_data.append([
                        row['First Name'],
                        row.get('Department', 'N/A'),
                        f"{row['Normal OT(H)']:.1f}",
                        level
                    ])
                
                ot_table = Table(ot_data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.3*inch])
                ot_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(ot_table)
            story.append(Spacer(1, 0.3*inch))
        
        story.append(PageBreak())
        
        # Action Items
        story.append(Paragraph("HR Action Items", heading_style))
        story.append(Paragraph("Employees requiring immediate attention:", styles['Normal']))
        story.append(Spacer(1, 0.1*inch))
        
        # Build action items list
        action_list = self._generate_action_items_list()
        for action in action_list[:15]:  # Top 15 items
            story.append(Paragraph(f"• {action}", styles['Normal']))
        
        story.append(Spacer(1, 0.3*inch))
        
        # Footer
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        return output_path
    
    def _generate_action_items_list(self):
        """Generate list of action items for PDF"""
        actions = []
        
        if 'Late In(M)' in self.df.columns:
            late_critical = self.df[self.df['Late In(M)'] > 100]
            for _, row in late_critical.iterrows():
                actions.append(
                    f"<b>{row['First Name']}</b> (ID {row['Employee ID']}): "
                    f"Excessive late arrivals ({row['Late In(M)']:.0f} min) - Schedule counseling"
                )
        
        if 'Normal OT(H)' in self.df.columns:
            ot_critical = self.df[self.df['Normal OT(H)'] > 30]
            for _, row in ot_critical.iterrows():
                actions.append(
                    f"<b>{row['First Name']}</b> (ID {row['Employee ID']}): "
                    f"High overtime ({row['Normal OT(H)']:.1f}h) - Review workload"
                )
        
        return actions if actions else ["No critical issues - all metrics within acceptable ranges"]
    
    def generate_both_reports(self, excel_path='attendance_report.xlsx', 
                             pdf_path='attendance_report.pdf'):
        """
        Generate both Excel and PDF reports
        
        Args:
            excel_path: Path for Excel file
            pdf_path: Path for PDF file
            
        Returns:
            tuple: (excel_path, pdf_path)
        """
        excel_file = self.generate_excel_report(excel_path)
        pdf_file = self.generate_pdf_report(pdf_path)
        
        return excel_file, pdf_file